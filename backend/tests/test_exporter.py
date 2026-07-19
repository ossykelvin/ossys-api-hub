from io import BytesIO

from openpyxl import load_workbook

from app.services.exporter import (
    create_csv,
    create_xlsx,
    flatten_record,
    flatten_records,
    tabular_export_records,
)
from app.services.paths import get_path, normalize_records


def test_get_path_and_normalize_records():
    payload = {"data": {"items": [{"node": {"id": 1}}, {"node": {"id": 2}}]}}
    items = get_path(payload, "data.items")
    assert normalize_records(items, "node") == [{"id": 1}, {"id": 2}]


def test_flatten_record():
    record = {"id": 1, "customer": {"name": "Ada"}, "tags": ["a", "b"]}
    flattened = flatten_record(record)
    assert flattened["customer.name"] == "Ada"
    assert flattened["tags"] == '["a", "b"]'


def test_flatten_records_preserves_column_order():
    columns, rows = flatten_records([{"id": 1, "a": 2}, {"id": 2, "b": 3}])
    assert columns == ["id", "a", "b"]
    assert rows[1]["b"] == 3


def test_csv_escapes_spreadsheet_formulas():
    content = create_csv([{"=name": "=HYPERLINK(\"bad\")", "count": -2}]).decode("utf-8-sig")
    assert content.startswith("'=name,count")
    assert "'=HYPERLINK" in content
    assert ",-2" in content


def test_xlsx_escapes_formulas_and_invalid_xml_characters():
    path = create_xlsx(
        [{"name": "@malicious\x00value"}],
        "query { value }",
        {},
        "https://example.com/graphql",
        {},
        [],
    )
    try:
        workbook = load_workbook(BytesIO(path.read_bytes()), data_only=False)
        assert workbook["Data"]["A2"].value == "'@maliciousvalue"
    finally:
        path.unlink(missing_ok=True)


def test_tabular_export_records_combines_nested_results_across_pages():
    records = [
        {"data": {"items": {"data": [{"id": 1}], "pageNumber": 1}}},
        {"data": {"items": {"data": [{"id": 2}], "pageNumber": 2}}},
    ]

    data_records, response_summary, result_path = tabular_export_records(records)

    assert data_records == [{"id": 1}, {"id": 2}]
    assert response_summary == records
    assert result_path == "data.items.data"


def test_tabular_export_records_keeps_existing_rows_with_nested_child_lists():
    records = [
        {"id": 1, "tags": [{"name": "one"}]},
        {"id": 2, "tags": [{"name": "two"}]},
    ]

    data_records, response_summary, result_path = tabular_export_records(records)

    assert data_records == records
    assert response_summary is None
    assert result_path is None


def test_xlsx_expands_nested_result_fields_and_preserves_response_summary():
    records = [{
        "data": {
            "sensorAssignments": {
                "data": [
                    {"id": "a1", "sensorId": "s1", "settings": {"reportsOnOccupancy": True}},
                    {"id": "a2", "sensorId": "s2", "settings": {"reportsOnOccupancy": False}},
                ],
                "count": 2,
                "total": 2,
                "pageNumber": 1,
                "pageSize": 100,
            },
        },
    }]
    path = create_xlsx(records, "query", {}, "https://example.com/graphql", {}, [])

    try:
        workbook = load_workbook(BytesIO(path.read_bytes()), data_only=False)
        data_sheet = workbook["Data"]
        assert [cell.value for cell in data_sheet[1]] == [
            "id",
            "sensorId",
            "settings.reportsOnOccupancy",
        ]
        assert [cell.value for cell in data_sheet[2]] == ["a1", "s1", True]
        assert [cell.value for cell in data_sheet[3]] == ["a2", "s2", False]

        response_sheet = workbook["Response Summary"]
        assert "data.sensorAssignments.data" in [cell.value for cell in response_sheet[1]]
        assert "data.sensorAssignments.count" in [cell.value for cell in response_sheet[1]]
        assert workbook["Run Summary"]["A2"].value == "Endpoint"
        assert any(
            row[0].value == "Expanded result path" and row[1].value == "data.sensorAssignments.data"
            for row in workbook["Run Summary"].iter_rows()
        )
    finally:
        path.unlink(missing_ok=True)
