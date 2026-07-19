from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_INVALID_FILENAME = re.compile(r"[^A-Za-z0-9._-]+")
_ILLEGAL_XLSX_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_RESULT_COLLECTION_KEYS = {"data", "edges", "items", "nodes", "records", "results", "value", "values"}


def safe_filename(value: str, extension: str) -> str:
    stem = _INVALID_FILENAME.sub("-", value.strip()).strip("-.") or "graphql-report"
    return f"{stem}.{extension}"


def safe_spreadsheet_value(value: Any) -> Any:
    """Keep untrusted text from becoming a formula or invalid workbook XML."""
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL_XLSX_CHARACTERS.sub("", value)
    if cleaned.lstrip().startswith(_FORMULA_PREFIXES):
        return f"'{cleaned}"
    return cleaned


def flatten_record(record: dict[str, Any], prefix: str = "", separator: str = ".") -> dict[str, Any]:
    output: dict[str, Any] = {}
    for key, value in record.items():
        column = f"{prefix}{separator}{key}" if prefix else str(key)
        if isinstance(value, dict):
            output.update(flatten_record(value, column, separator))
        elif isinstance(value, list):
            output[column] = json.dumps(value, ensure_ascii=False, default=str)
        else:
            output[column] = value
    return output


def flatten_records(records: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    flattened = [flatten_record(record) for record in records]
    columns: list[str] = []
    seen: set[str] = set()
    for row in flattened:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns, flattened


def _nested_record_arrays(value: Any, path: tuple[str, ...] = ()) -> list[tuple[tuple[str, ...], list[dict[str, Any]]]]:
    candidates: list[tuple[tuple[str, ...], list[dict[str, Any]]]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            candidates.extend(_nested_record_arrays(child, (*path, str(key))))
    elif isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
        candidates.append((path, value))
    return candidates


def tabular_export_records(
    records: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]] | None, str | None]:
    """Extract a nested result collection while retaining its response envelopes."""
    records_look_tabular = len(records) > 1 and any(
        any(not isinstance(value, (dict, list)) for value in record.values())
        for record in records
    )
    if records_look_tabular:
        return records, None, None

    candidates_by_path: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    for record in records:
        for path, nested_records in _nested_record_arrays(record):
            has_root_scalar = any(
                not isinstance(value, (dict, list))
                for value in record.values()
            )
            if has_root_scalar and path[-1].casefold() not in _RESULT_COLLECTION_KEYS:
                continue
            candidates_by_path.setdefault(path, []).extend(nested_records)

    if not candidates_by_path:
        return records, None, None

    selected_path, selected_records = max(
        candidates_by_path.items(),
        key=lambda candidate: (len(candidate[1]), len(candidate[0])),
    )
    return selected_records, records, ".".join(selected_path)


def create_csv(records: list[dict[str, Any]]) -> bytes:
    columns, rows = flatten_records(records)
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow([safe_spreadsheet_value(column) for column in columns])
    writer.writerows(
        [safe_spreadsheet_value(row.get(column)) for column in columns]
        for row in rows
    )
    return stream.getvalue().encode("utf-8-sig")


def create_json(records: list[dict[str, Any]]) -> bytes:
    return json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8")


def create_xlsx(
    records: list[dict[str, Any]],
    query: str,
    variables: dict[str, Any],
    endpoint: str,
    run_summary: dict[str, Any],
    errors: list[dict[str, Any]],
) -> Path:
    data_records, response_summary_records, result_path = tabular_export_records(records)
    columns, rows = flatten_records(data_records)
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)

    if columns:
        data_sheet.append([safe_spreadsheet_value(column) for column in columns])
        for cell in data_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        data_sheet.freeze_panes = "A2"
        data_sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"
        for row in rows:
            data_sheet.append([safe_spreadsheet_value(row.get(column)) for column in columns])
        sample_rows = rows[:200]
        for index, column in enumerate(columns, start=1):
            max_length = max([len(column)] + [len(str(row.get(column, ""))) for row in sample_rows])
            data_sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 55)
    else:
        data_sheet.append(["No records returned"])

    if response_summary_records is not None:
        summary_columns, summary_rows = flatten_records(response_summary_records)
        response_sheet = workbook.create_sheet("Response Summary")
        response_sheet.append([safe_spreadsheet_value(column) for column in summary_columns])
        for cell in response_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
        response_sheet.freeze_panes = "A2"
        response_sheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(summary_columns))}{max(1, len(summary_rows) + 1)}"
        )
        for row in summary_rows:
            response_sheet.append([
                safe_spreadsheet_value(row.get(column))
                for column in summary_columns
            ])
        for index, column in enumerate(summary_columns, start=1):
            sample_values = [len(str(row.get(column, ""))) for row in summary_rows[:200]]
            max_length = max([len(column), *sample_values])
            response_sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 55)

    summary_sheet = workbook.create_sheet("Run Summary")
    summary_sheet.append(["Metric", "Value"])
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary_values = {"Endpoint": endpoint, **run_summary}
    if result_path is not None:
        summary_values["Expanded result path"] = result_path
    for key, value in summary_values.items():
        display_value = json.dumps(value, ensure_ascii=False, default=str) if isinstance(value, (dict, list)) else value
        summary_sheet.append([safe_spreadsheet_value(str(key)), safe_spreadsheet_value(display_value)])
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 80

    query_sheet = workbook.create_sheet("Query")
    query_sheet.append(["GraphQL Query"])
    query_sheet["A1"].fill = header_fill
    query_sheet["A1"].font = header_font
    query_sheet.append([safe_spreadsheet_value(query)])
    query_sheet.append([])
    query_sheet.append(["Variables"])
    query_sheet["A4"].fill = header_fill
    query_sheet["A4"].font = header_font
    query_sheet.append([safe_spreadsheet_value(json.dumps(variables, ensure_ascii=False, indent=2, default=str))])
    query_sheet.column_dimensions["A"].width = 120
    query_sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    query_sheet["A5"].alignment = Alignment(wrap_text=True, vertical="top")

    error_sheet = workbook.create_sheet("Errors")
    error_sheet.append(["Page", "Message", "Details"])
    for cell in error_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for error in errors:
        error_sheet.append([
            error.get("page"),
            safe_spreadsheet_value(error.get("message")),
            safe_spreadsheet_value(json.dumps(error.get("details"), ensure_ascii=False, default=str)) if error.get("details") is not None else "",
        ])
    error_sheet.column_dimensions["A"].width = 12
    error_sheet.column_dimensions["B"].width = 45
    error_sheet.column_dimensions["C"].width = 80

    temp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp.close()
    workbook.save(temp.name)
    return Path(temp.name)
